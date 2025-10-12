from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.pagination import PageNumberPagination
from django.conf import settings
from rest_framework import viewsets
from rest_framework.views import APIView
from .models import (
    Run,
    StatusChoices,
    AthleteInfo,
    Challenge,
    Position,
    CollectibleItem,
    Subscribe,
)
from .serializers import (
    RunSerializer,
    UserSerializer,
    ChallengesSerializer,
    PositionSerializer,
    CollectibleItemSerializer,
    AthleteSerializerExtended,
    CoachSerializerExtended,
    TotalChallengesSerializer,
)
from django.contrib.auth.models import User
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Q, Sum, Max, Min, Avg
from geopy.distance import geodesic
from openpyxl import load_workbook
from .utils import validate_latitude, validate_longitude
from datetime import timedelta
from collections import defaultdict


@api_view(["GET"])
def get_company_details(request):
    """
    Sends main information about the company.
    """
    return Response(settings.COMPANY_INFORMATION)


class AppPagination(PageNumberPagination):
    page_size_query_param = "size"
    max_page_size = 50


class RunViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing and editing run instances.
    There is also additional fetch for User Model via select_related through
    athlete field.
    """

    queryset = Run.objects.select_related("athlete").all()
    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["status", "athlete"]
    ordering_fields = ["created_at"]
    ordering = ["id"]
    pagination_class = AppPagination


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    """
    A viewset is for read only. It allows to see users.
    It accepts Query Parameters:
    ?type=coach
    ?type=athlete
    ?type=*
    Coach returns users when is_staff property is True
    Athlete returns users where is_staff is False
    if another argument is provided viewSet returns both
    The viewSet does not show superusers.

    Additionally viewSet allows to search users by first_name
    or last_name.
    """

    queryset = User.objects.exclude(is_superuser=True).annotate(
        runs_finished_count=Count(
            "run",
            filter=Q(run__status=StatusChoices.FINISHED),
        ),
        avg_rating=Avg("coach__rating"),
    )
    serializer_class = UserSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["first_name", "last_name"]
    ordering = ["id"]
    ordering_fields = ["date_joined"]
    pagination_class = AppPagination

    def get_queryset(self):
        param_type = self.request.query_params.get("type")
        if param_type == "coach":
            queryset = self.queryset.filter(is_staff=True)
        elif param_type == "athlete":
            queryset = self.queryset.filter(is_staff=False)
        else:
            queryset = self.queryset
        return queryset

    def get_serializer_class(self):
        if "pk" in self.kwargs:
            user = self.get_object()
            if user.is_staff:
                return CoachSerializerExtended
            return AthleteSerializerExtended
        return UserSerializer


class StartRunView(APIView):
    """
    Change status for Run instance to IN_PROGRESS
    Will return 404 if no object found
    Will raise 400 bad request if run status is IN_PROGRESS or FINISHED
    """

    def post(self, request, id):
        run = get_object_or_404(Run, id=id)
        if (
            run.status == StatusChoices.IN_PROGRESS
            or run.status == StatusChoices.FINISHED
        ):
            data = {"status": "bad_request"}
            return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)
        run.status = StatusChoices.IN_PROGRESS
        run.save()
        data = {"status": "success"}
        return JsonResponse(data, status=status.HTTP_200_OK)


class StopRunView(APIView):
    """
    Change status for Run instance to FINISHED
    Will return 404 if no object found
    Will raise 400 bad request if run instance has not been started.
    If two positions are related to current run it will calculate distance
    """

    challenge_name_10_runs = "Сделай 10 Забегов!"
    challenge_name_50_km = "Пробеги 50 километров!"
    challenge_name_2_km = "2 километра за 10 минут!"

    def post(self, request, id):
        run = get_object_or_404(Run.objects.select_related("athlete"), id=id)
        # Check that run status is not INIT nor FINISHED
        self.check_correct_status(run)
        run.status = StatusChoices.FINISHED
        # get total km and run_time_seconds
        distance, result_time = self.get_total_km(run)
        if not self.has_challenge(
            run.athlete, self.challenge_name_2_km
        ) and self.check_2km_10min(distance, result_time):
            Challenge.objects.create(
                athlete=run.athlete, full_name=self.challenge_name_2_km
            )

        if not self.has_challenge(
            run.athlete, self.challenge_name_10_runs
        ) and self.has_ten_runs(run.athlete):
            Challenge.objects.create(
                athlete=run.athlete, full_name=self.challenge_name_10_runs
            )
        self.calculate_total_distance(run.athlete)
        data = {"status": "success"}
        return JsonResponse(data, status=status.HTTP_200_OK)

    def has_ten_runs(self, user):
        runs_count = Run.objects.filter(
            athlete=user, status=StatusChoices.FINISHED
        ).count()
        return runs_count >= 10

    def has_challenge(self, user, challenge_name):
        return Challenge.objects.filter(athlete=user, full_name=challenge_name).exists()

    def calculate_total_distance(self, user):
        total_distance = Run.objects.filter(athlete=user).aggregate(Sum("distance"))
        if (
            not self.has_challenge(user, self.challenge_name_50_km)
            and total_distance["distance__sum"]
            and total_distance["distance__sum"] >= 50
        ):
            Challenge.objects.create(athlete=user, full_name=self.challenge_name_50_km)
        return

    def check_correct_status(self, run):
        if run.status == StatusChoices.INIT or run.status == StatusChoices.FINISHED:
            data = {"status": "bad_request"}
            return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    def get_total_km(self, run):
        positions = Position.objects.filter(run=run)
        length = len(positions)
        total = 0
        result_time = timedelta(seconds=0)
        if length > 1:
            for index, position in enumerate(positions):
                if index == length - 1:
                    break
                start = (positions[index].latitude, positions[index].longitude)
                finish = (positions[index + 1].latitude, positions[index + 1].longitude)
                total += geodesic(start, finish).km
            run.distance = round(total, 3)
            # find max and min time
            date_time_stats = positions.aggregate(
                min_time=Min("date_time"), max_time=Max("date_time")
            )
            result_time = date_time_stats["max_time"] - date_time_stats["min_time"]
            run.run_time_seconds = round(result_time.total_seconds())

            # avarage speed (meters per seconds)
            avg_speed = positions.aggregate(avg_speed=Avg("speed"))["avg_speed"]
            run.speed = round(avg_speed, 2) if avg_speed is not None else 0
        else:
            run.run_time_seconds = 0
            run.distance = 0
            run.speed = 0
        run.save()
        return total, result_time

    def check_2km_10min(self, distance, result_time_sec):
        ten_minutes = timedelta(minutes=10)
        return distance >= 2 and result_time_sec <= ten_minutes


class AthleteInfoView(APIView):
    """
    GET or PUT request to see additional info about athlete
    If AthleteInfo does not exist django will create it.
    """

    def get(self, request, id):
        athlete, _ = AthleteInfo.objects.select_related("user_id").get_or_create(
            user_id_id=id
        )
        data = {
            "goals": athlete.goals,
            "weight": athlete.weight,
            "user_id": athlete.user_id.id,
        }
        return JsonResponse(data, status=status.HTTP_200_OK)

    def put(self, request, id):
        goals = request.data.get("goals")
        weight = request.data.get("weight")

        if not goals or not weight:
            return JsonResponse(
                {"detail": "Please provide goals and weight"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            weight = int(weight)
            assert 0 < weight < 900
        except (ValueError, AssertionError):
            return JsonResponse(
                {"detail": "weight should be greater than 0 and smaller than 900"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        athlete, _ = AthleteInfo.objects.update_or_create(
            user_id_id=id,
            defaults={"goals": goals, "weight": weight},
        )

        return JsonResponse(
            {"goals": athlete.goals, "weight": athlete.weight, "user_id": id},
            status=status.HTTP_201_CREATED,
        )


class ChallengesViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Shows all challenges by athletes
    If params provided /?athlete=<id> this endpoint retrieves challenges by a particular athlete
    """

    serializer_class = ChallengesSerializer

    def get_queryset(self):
        queryset = Challenge.objects.all()
        athlete_id = self.request.query_params.get("athlete")
        if athlete_id:
            queryset = queryset.filter(athlete_id=athlete_id)
        return queryset


class PositionViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing and editing position instances.
    There is validation in the serializer.
    If the position of the user is close to collectible items
    then we add this item to the user as a reward.
    """

    queryset = Position.objects.select_related("run", "run__athlete").all()
    serializer_class = PositionSerializer

    def get_queryset(self):
        param_type = self.request.query_params.get("run")
        try:
            param_type = int(param_type)
        except:
            param_type = ""

        if param_type:
            queryset = self.queryset.filter(run=param_type)
        else:
            queryset = self.queryset
        return queryset

    def perform_create(self, serializer):
        # Before save we find last position
        last_position = (
            Position.objects.filter(run=serializer.validated_data["run"])
            .order_by("-date_time")
            .first()
        )

        instance = serializer.save()
        current_position = (instance.latitude, instance.longitude)
        athlete = instance.run.athlete
        # это все бы надо в celery!
        self.check_collectible_awards(instance, current_position, athlete)
        self.check_speed(instance, last_position)

    def check_collectible_awards(self, instance, current_position, athlete):
        # radius of search is 0.1 km. Approx 0.001 degree
        delta = 0.001
        lat, lon = instance.latitude, instance.longitude

        # find collectible items that are close to the position of the runner
        collectible_items = CollectibleItem.objects.filter(
            latitude__range=(lat - delta, lat + delta),
            longitude__range=(lon - delta, lon + delta),
        )

        for item in collectible_items:
            if validate_latitude(item.latitude) and validate_longitude(item.longitude):
                desirable = (item.latitude, item.longitude)
                difference = geodesic(current_position, desirable).km
                if difference < 0.1:
                    item.users.add(athlete)

    def check_speed(self, instance, last_position_obj):
        if not last_position_obj:
            instance.speed = 0
            instance.distance = 0
            instance.save()
            return 0
        # calc speed v = s / t
        # s – distance, а t – time
        # our task is to get in meters per seconds
        last_position = (last_position_obj.latitude, last_position_obj.longitude)
        current_position = (instance.latitude, instance.longitude)
        distance_m = geodesic(current_position, last_position).meters
        # time delta
        delta_t = (instance.date_time - last_position_obj.date_time).total_seconds()
        # speed
        speed = distance_m / delta_t if delta_t > 0 else 0
        instance.speed = round(speed, 2)
        instance.distance = last_position_obj.distance + round(distance_m / 1000, 2)
        instance.save(update_fields=["distance", "speed"])


class CollectibleItemViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Shows all Collectible Items
    """

    serializer_class = CollectibleItemSerializer
    queryset = CollectibleItem.objects.all()


@api_view(["POST"])
def upload_collectible_items(request):
    """
    Allows to upload data from xlsx format
    """
    file = request.FILES.get("file")
    if not file:
        data = {"error": "Please provide xlsx file"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)
    # Проверка и чтение файла
    workbook = load_workbook(filename=file, data_only=True)
    sheet = workbook.active

    # Headers of the file
    headers = [
        "name",
        "uid",
        "value",
        "latitude",
        "longitude",
        "picture",
    ]

    errors = []

    for _, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # row is tuple now => make it dict
        item_data = dict(zip(headers, row))

        serializer = CollectibleItemSerializer(data=item_data)
        if serializer.is_valid():
            serializer.save()
        else:
            # [] for errors inside this row
            error_colums = []
            for field in serializer.fields:
                error_colums.append(item_data[field])
            errors.append(error_colums)

    return JsonResponse(errors, status=status.HTTP_200_OK, safe=False)


@api_view(["POST"])
def subscribe_to_coach(request, id):
    coach = get_object_or_404(User, pk=id)
    if not coach.is_staff:
        data = {"info": "Можно подписаться только на тренера"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    athlete_id = request.data.get("athlete")
    try:
        athlete = User.objects.get(pk=athlete_id)
    except User.DoesNotExist:
        data = {"info": f"Атлет с ID {athlete_id} не найден"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    if athlete.is_staff:
        data = {"info": "На тренера могут подписаться только бегуны"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    if Subscribe.objects.filter(coach=coach, athlete=athlete).exists():
        data = {"info": "Подписку можно оформить только 1 раз"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    subscription = Subscribe.objects.create(coach=coach, athlete=athlete)

    data = {"Подписка": subscription.id}

    return JsonResponse(data, status=status.HTTP_200_OK)


class ChallengesListView(APIView):

    def get(self, request):
        challenges_list = list(Challenge.objects.select_related("athlete").all())
        challenges_map = defaultdict(list)

        for challenge in challenges_list:
            if challenge.athlete:
                challenges_map[challenge.full_name].append(challenge.athlete)

        data = [
            {"name_to_display": name, "athletes": athletes}
            for name, athletes in challenges_map.items()
        ]

        serializer = TotalChallengesSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["POST"])
def rate_coach(request, coach_id):
    coach = get_object_or_404(User, pk=coach_id)
    if not coach.is_staff:
        data = {"info": "Можно дать оценку только тренеру"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    athlete_id = request.data.get("athlete")
    try:
        athlete = User.objects.get(pk=athlete_id)
    except User.DoesNotExist:
        data = {"info": f"Атлет с ID {athlete_id} не найден"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    try:
        rating = int(request.data.get("rating"))
    except:
        data = {"info": "Рейтинг не является числовым значением"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    if rating > 5 or rating < 1:
        data = {"info": f"Рейтинг {rating} должен быть от 1 до 5"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    if athlete.is_staff:
        data = {"info": "Дать оценку тренерам могут только бегуны"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    if not Subscribe.objects.filter(coach=coach, athlete=athlete).exists():
        data = {"info": "Дать оценку тренерам может атлет, который на него подписан"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    subscription = Subscribe.objects.get(coach=coach, athlete=athlete)
    subscription.rating = rating
    subscription.save()

    data = {"Новый рейтинг": rating}

    return JsonResponse(data, status=status.HTTP_200_OK)


@api_view(["GET"])
def analytics_for_coach(request, coach_id):
    coach = get_object_or_404(User, pk=coach_id)
    if not coach.is_staff:
        data = {"info": "Статистику можно получить только по тренеру"}
        return JsonResponse(data, status=status.HTTP_400_BAD_REQUEST)

    # 'longest_run_user': ...  # Id Бегуна который сделал самый длинный забег у этого Тренера
    # 'longest_run_value': ... # Дистанция самого длинного забега
    # longest_run = (
    #     Run.objects.select_related("athlete")
    #     .filter(athlete__athlete__coach=coach, distance__isnull=False)
    #     .order_by("-distance")
    #     .first()
    # )

    # 'total_run_user': ...    # Id Бегуна который пробежал в сумме больше всех у этого Тренера
    # 'total_run_value': ...   # Дистанция которую в сумме пробежал этот Бегун
    # total_run = (
    #     Run.objects.select_related("athlete")
    #     .filter(athlete__athlete__coach=coach, distance__isnull=False)
    #     .annotate(total_distance=Sum("distance"))
    #     .order_by("-total_distance")
    #     .first()
    # )

    # 'speed_avg_user': ...    #  Id Бегуна который всреднем бежал быстрее всех
    # 'speed_avg_value': ...   # Средняя скорость этого Бегуна
    # speed_avg = (
    #     Run.objects.select_related("athlete")
    #     .filter(athlete__athlete__coach=coach, speed__isnull=False)
    #     .annotate(avg_speed=Avg("speed"))
    #     .order_by("-avg_speed")
    #     .first()
    # )

    # if not (longest_run or total_run or speed_avg):
    #     return JsonResponse(
    #         {"info": "У этого тренера пока нет забегов у атлетов"}, status=200
    #     )

    # Берём только атлетов, подписанных на этого тренера
    athletes = User.objects.filter(athlete__coach=coach)

    if not athletes.exists():
        return JsonResponse(
            {"info": "У этого тренера пока нет атлетов"},
            status=status.HTTP_200_OK,
        )

    athletes = (
        athletes.annotate(
            longest_run_value=Max("run__distance"),
            total_run_value=Sum("run__distance"),
            speed_avg_value=Avg("run__speed"),
        )
        .filter(run__distance__isnull=False)
        .distinct()
    )

    if not athletes.exists():
        return JsonResponse(
            {"info": "У этого тренера пока нет забегов у атлетов"},
            status=status.HTTP_200_OK,
        )

    longest_run_user = max(athletes, key=lambda u: u.longest_run_value or 0)

    total_run_user = max(athletes, key=lambda u: u.total_run_value or 0)

    speed_avg_user = max(athletes, key=lambda u: u.speed_avg_value or 0)

    # data = {
    #     "longest_run_user": longest_run.athlete_id,
    #     "longest_run_value": longest_run.distance,
    #     "total_run_user": total_run.athlete_id,
    #     "total_run_value": total_run.total_distance,
    #     "speed_avg_user": speed_avg.athlete_id,
    #     "speed_avg_value": speed_avg.avg_speed,
    # }

    data = {
        "longest_run_user": longest_run_user.id,
        "longest_run_value": round(longest_run_user.longest_run_value or 0, 3),
        "total_run_user": total_run_user.id,
        "total_run_value": round(total_run_user.total_run_value or 0, 3),
        "speed_avg_user": speed_avg_user.id,
        "speed_avg_value": round(speed_avg_user.speed_avg_value or 0, 2),
    }

    return JsonResponse(data, status=status.HTTP_200_OK)
